#!/usr/bin/env python3
"""Generate a formatted XLSX estimation workbook from JSON data.

Supports dynamic per-platform columns: when the JSON includes a "platforms"
array, the Estimates sheet gains one column per platform between "So that..."
and "Points", and an extra "Platform Summary" sheet is created.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
UNCERTAINTY_COLORS = {
    "Low": "C6EFCE",
    "Medium": "FFEB9C",
    "High": "FFC7CE",
}

_BASE_HEADERS = [
    "#",
    "Category",
    "Epic",
    "Area",
    "Story",
    "Role",
    "I want...",
    "So that...",
]
_POINTS_COL = "Points"
_RATIONALE_COL = "Rationale"
_UNCERTAINTY_COL = "Uncertainty"
# platform columns are inserted between _BASE_HEADERS and _POINTS_COL


def _style_header(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER


def _auto_width(ws, max_width: int = 50) -> None:
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                for line in str(cell.value).split("\n"):
                    lengths.append(len(line))
        if lengths:
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(lengths) + 3, max_width)


def _col_idx_for(field_name: str, platforms: list[str]) -> int:
    """Return 1-based column index for a named field."""
    mapping = {h: idx for idx, h in enumerate(field_order(platforms), 1)}
    return mapping[field_name]


def field_order(platforms: list[str]) -> list[str]:
    """Full header list for the Estimates sheet, with platforms inserted."""
    headers = list(_BASE_HEADERS)
    headers.extend(platforms)
    headers.append(_POINTS_COL)
    headers.append(_RATIONALE_COL)
    headers.append(_UNCERTAINTY_COL)
    return headers


def _build_estimates_sheet(
    wb: Workbook, estimates: list[dict], platforms: list[str]
) -> None:
    ws = wb.active
    ws.title = "Estimates"

    headers = field_order(platforms)
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, est in enumerate(estimates, 1):
        row_data = [
            i,
            est.get("category", ""),
            est.get("epic", ""),
            est.get("area", ""),
            est.get("story", ""),
            est.get("role", ""),
            est.get("want", ""),
            est.get("so_that", ""),
        ]
        pp = est.get("platform_points", {})
        for plat in platforms:
            row_data.append(pp.get(plat, 0))
        row_data.append(est.get("points", 0))
        row_data.append(est.get("rationale", ""))
        row_data.append(est.get("uncertainty", ""))

        ws.append(row_data)
        row_num = i + 1

        points_col = _col_idx_for(_POINTS_COL, platforms)
        unc_col = _col_idx_for(_UNCERTAINTY_COL, platforms)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Bold Points column
        ws.cell(row=row_num, column=points_col).font = Font(bold=True, size=11)
        ws.cell(row=row_num, column=points_col).alignment = Alignment(
            horizontal="center", vertical="top"
        )

        # Center-align platform columns
        for j, _ in enumerate(platforms, 1):
            plat_col = _col_idx_for(platforms[j - 1], platforms)
            ws.cell(row=row_num, column=plat_col).alignment = Alignment(
                horizontal="center", vertical="top"
            )

        # Color uncertainty column
        uncertainty = est.get("uncertainty", "")
        fill_color = UNCERTAINTY_COLORS.get(uncertainty)
        if fill_color:
            ws.cell(row=row_num, column=unc_col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            ws.cell(row=row_num, column=unc_col).alignment = Alignment(
                horizontal="center", vertical="top"
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws)


def _build_platform_summary(
    wb: Workbook,
    estimates: list[dict],
    platforms: list[str],
    functional_total: int,
    cfr_total: int,
    grand_total: int,
    buffers: list[dict],
) -> None:
    ws = wb.create_sheet("Platform Summary")

    # --- Per-platform totals ---
    row = 1
    ws.cell(row=row, column=1, value="Module").font = Font(
        bold=True, size=11, color="2F5496"
    )
    for j, plat in enumerate(platforms, 2):
        ws.cell(row=row, column=j, value=plat).font = Font(
            bold=True, size=11, color="2F5496"
        )
    total_col = len(platforms) + 2
    ws.cell(row=row, column=total_col, value="Subtotal (SP)").font = Font(
        bold=True, size=11, color="2F5496"
    )
    for col in range(1, total_col + 1):
        ws.cell(row=row, column=col).fill = HEADER_FILL
        ws.cell(row=row, column=col).font = HEADER_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

    row = 2

    def _row_data(label: str, estimate_filter, is_total: bool = False):
        nonlocal row
        subset = [e for e in estimates if estimate_filter(e)]
        vals = []
        row_subtotal = 0
        for plat in platforms:
            plat_sum = sum(e.get("platform_points", {}).get(plat, 0) for e in subset)
            vals.append(plat_sum)
            row_subtotal += plat_sum
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        for j, v in enumerate(vals, 2):
            c = ws.cell(row=row, column=j, value=v if v else ("—" if not is_total else v))
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
        st = ws.cell(row=row, column=total_col, value=row_subtotal)
        st.border = THIN_BORDER
        st.alignment = Alignment(horizontal="center")
        if is_total:
            st.font = TOTAL_FONT
        else:
            st.font = Font(bold=True)
        row += 1
        return row_subtotal

    func_sub = _row_data("Functional Stories", lambda e: e.get("category") == "Functional")
    cfr_sub = _row_data("CFR Items", lambda e: e.get("category") == "CFR")
    func_cfr_total = func_sub + cfr_sub
    _row_data("Functional + CFR Total", lambda _: True, is_total=True)

    row += 1
    if buffers:
        buf_total = 0
        for buf in buffers:
            pts = buf.get("points", 0)
            label = f"Buffer: {buf.get('type', '')}  ({buf.get('pct', '')})"
            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            ws.cell(row=row, column=1).font = Font(bold=True)
            bc = ws.cell(row=row, column=total_col, value=pts)
            bc.border = THIN_BORDER
            bc.alignment = Alignment(horizontal="center")
            bc.font = Font(bold=True)
            buf_total += pts
            row += 1
            # Add rationale on the next line
            if buf.get("rationale"):
                ws.cell(row=row, column=1, value=f"     {buf['rationale']}").font = Font(
                    italic=True, size=10, color="666666"
                )
                row += 1

        ws.cell(row=row, column=1, value="Buffer Total").border = THIN_BORDER
        ws.cell(row=row, column=1).font = TOTAL_FONT
        bt = ws.cell(row=row, column=total_col, value=buf_total)
        bt.border = THIN_BORDER
        bt.alignment = Alignment(horizontal="center")
        bt.font = TOTAL_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Recommended Planning Estimate").border = THIN_BORDER
    ws.cell(row=row, column=1).font = TOTAL_FONT
    total_with_buffer = func_cfr_total + sum(b.get("points", 0) for b in buffers)
    rpe = ws.cell(row=row, column=total_col, value=total_with_buffer)
    rpe.border = THIN_BORDER
    rpe.alignment = Alignment(horizontal="center")
    rpe.font = TOTAL_FONT

    ws.column_dimensions["A"].width = 28
    _auto_width(ws, max_width=30)


def _build_raid_sheet(wb: Workbook, raid: list[dict]) -> None:
    ws = wb.create_sheet("RAID")

    headers = ["Type", "Item", "Impact", "Mitigation"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, item in enumerate(raid, 1):
        row_data = [
            item.get("type", ""),
            item.get("item", ""),
            item.get("impact", ""),
            item.get("mitigation", ""),
        ]
        ws.append(row_data)
        row_num = i + 1
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = THIN_BORDER
            ws.cell(row=row_num, column=col).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).alignment = Alignment(
            horizontal="center", vertical="top"
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_width(ws, max_width=60)


def _build_summary_sheet(
    wb: Workbook,
    data: dict,
    functional_total: int,
    cfr_total: int,
    grand_total: int,
) -> None:
    ws = wb.create_sheet("Summary")

    row = 1
    ws.cell(row=row, column=1, value="Scope Summary").font = Font(
        bold=True, size=14, color="2F5496"
    )
    row += 1
    ws.cell(row=row, column=1, value=data.get("scope_summary", "")).alignment = Alignment(
        wrap_text=True
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # --- Category Breakdown ---
    row += 2
    ws.cell(row=row, column=1, value="Category Breakdown").font = Font(
        bold=True, size=14, color="2F5496"
    )
    row += 1
    cat_headers = ["Category", "Total Points", "%"]
    for ci, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for cat_name, cat_pts in [
        ("Functional Stories", functional_total),
        ("Non-Functional (CFR)", cfr_total),
    ]:
        ws.cell(row=row, column=1, value=cat_name).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        pt_cell = ws.cell(row=row, column=2, value=cat_pts)
        pt_cell.border = THIN_BORDER
        pt_cell.alignment = Alignment(horizontal="center")
        pt_cell.font = Font(bold=True)
        pct = (cat_pts / grand_total * 100) if grand_total > 0 else 0
        pct_cell = ws.cell(row=row, column=3, value=f"{pct:.0f}%")
        pct_cell.border = THIN_BORDER
        pct_cell.alignment = Alignment(horizontal="center")
        row += 1

    ws.cell(row=row, column=1, value="Grand Total").border = THIN_BORDER
    ws.cell(row=row, column=1).font = TOTAL_FONT
    gt_cell = ws.cell(row=row, column=2, value=grand_total)
    gt_cell.border = THIN_BORDER
    gt_cell.font = TOTAL_FONT
    gt_cell.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value="100%").border = THIN_BORDER
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3).font = TOTAL_FONT

    # --- Confidence Analysis ---
    row += 2
    ws.cell(row=row, column=1, value="Confidence Analysis").font = Font(
        bold=True, size=14, color="2F5496"
    )
    row += 1
    ws.cell(
        row=row, column=1,
        value=f"Overall Confidence: {data.get('confidence', 'N/A')}",
    ).font = Font(bold=True, size=11)

    row += 1
    unc_headers = ["Uncertainty Level", "Points", "% of Total"]
    for ci, h in enumerate(unc_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    uncertainty_counts: dict[str, int] = {"Low": 0, "Medium": 0, "High": 0}
    for est in data.get("estimates", []):
        level = est.get("uncertainty", "")
        if level in uncertainty_counts:
            uncertainty_counts[level] += est.get("points", 0)

    row += 1
    for level in ("Low", "Medium", "High"):
        pts = uncertainty_counts[level]
        ws.cell(row=row, column=1, value=level).border = THIN_BORDER
        ws.cell(row=row, column=1).font = Font(bold=True)
        fill_color = UNCERTAINTY_COLORS.get(level)
        if fill_color:
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
        pt_cell = ws.cell(row=row, column=2, value=pts)
        pt_cell.border = THIN_BORDER
        pt_cell.alignment = Alignment(horizontal="center")
        pt_cell.font = Font(bold=True)
        pct = (pts / grand_total * 100) if grand_total > 0 else 0
        pct_cell = ws.cell(row=row, column=3, value=f"{pct:.0f}%")
        pct_cell.border = THIN_BORDER
        pct_cell.alignment = Alignment(horizontal="center")
        row += 1

    # --- Key Assumptions ---
    row += 1
    ws.cell(row=row, column=1, value="Key Assumptions").font = Font(
        bold=True, size=14, color="2F5496"
    )
    row += 1
    for assumption in data.get("key_assumptions", []):
        ws.cell(row=row, column=1, value=f"  {assumption}").alignment = Alignment(
            wrap_text=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    # --- Recommendations ---
    if data.get("recommendations"):
        row += 1
        ws.cell(row=row, column=1, value="Recommendations").font = Font(
            bold=True, size=14, color="2F5496"
        )
        row += 1
        ws.cell(row=row, column=1, value=data["recommendations"]).alignment = Alignment(
            wrap_text=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate a formatted XLSX estimation workbook."
    )
    parser.add_argument("--output", "-o", required=True, help="Output XLSX file path")
    parser.add_argument(
        "--data", "-d", required=True, help="JSON data file with estimation data"
    )
    args = parser.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"Error: data file not found: {args.data}", file=sys.stderr)
        sys.exit(1)

    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)

    estimates: list[dict] = data.get("estimates", [])
    raid: list[dict] = data.get("raid", [])
    platforms: list[str] = data.get("platforms", [])
    buffers: list[dict] = data.get("buffers", [])

    functional_total = sum(
        e.get("points", 0) for e in estimates if e.get("category") == "Functional"
    )
    cfr_total = sum(
        e.get("points", 0) for e in estimates if e.get("category") == "CFR"
    )
    grand_total = functional_total + cfr_total

    wb = Workbook()
    _build_estimates_sheet(wb, estimates, platforms)
    if platforms:
        _build_platform_summary(
            wb, estimates, platforms,
            functional_total, cfr_total, grand_total, buffers,
        )
    _build_raid_sheet(wb, raid)
    _build_summary_sheet(wb, data, functional_total, cfr_total, grand_total)

    output_path = Path(args.output)
    wb.save(output_path)
    print(f"Estimation workbook saved to: {output_path.resolve()}")


if __name__ == "__main__":
    main()
