#!/usr/bin/env python3
"""Excel file operations CLI for AI coding assistants.

Provides subcommands for reading, writing, searching, and manipulating
Excel workbooks (.xlsx) via openpyxl. Legacy .xls files are auto-converted
to .xlsx via xls2xlsx if available.

Requires: openpyxl (pip install openpyxl)
Optional: xls2xlsx (pip install xls2xlsx) — for .xls support
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print(
        "ERROR: openpyxl is not installed.\n"
        "Install it with: pip install openpyxl\n"
        "Or: pip install mythril-agent-skills",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _is_xls(path: str) -> bool:
    """Check if the file is a legacy .xls format."""
    return Path(path).suffix.lower() == ".xls"


def _convert_xls_to_xlsx(xls_path: str) -> str:
    """Convert a .xls file to .xlsx using xls2xlsx. Returns the .xlsx path.

    The converted file is placed next to the original with .xlsx extension.
    """
    try:
        from xls2xlsx import XLS2XLSX
    except ImportError:
        print(
            "ERROR: Cannot read .xls files without xls2xlsx.\n"
            "Install it with: pip install xls2xlsx\n"
            "Or convert the file to .xlsx first (e.g. via LibreOffice).",
            file=sys.stderr,
        )
        sys.exit(2)

    p = Path(xls_path)
    xlsx_path = p.with_suffix(".xlsx")
    if xlsx_path.exists():
        print(f"Using existing converted file: {xlsx_path}")
        return str(xlsx_path)

    try:
        x2x = XLS2XLSX(str(p))
        x2x.to_xlsx(str(xlsx_path))
        print(f"Converted: {p.name} -> {xlsx_path.name}")
        return str(xlsx_path)
    except Exception as exc:
        print(f"ERROR: Failed to convert .xls to .xlsx: {exc}", file=sys.stderr)
        sys.exit(2)


def _load_workbook(path: str, *, data_only: bool = False) -> openpyxl.Workbook:
    """Load a workbook, handling .xls auto-conversion and common errors."""
    p = Path(path)
    if not p.exists():
        print(f"ERROR: File not found: {path}", file=sys.stderr)
        sys.exit(2)

    actual_path = path
    if _is_xls(path):
        actual_path = _convert_xls_to_xlsx(path)

    try:
        return openpyxl.load_workbook(actual_path, data_only=data_only)
    except Exception as exc:
        print(f"ERROR: Cannot open workbook: {exc}", file=sys.stderr)
        sys.exit(2)


def _output_path(source: str, in_place: bool) -> str:
    """Determine save path: in-place or timestamped copy next to the source.

    For .xls source files, always output .xlsx (can't write .xls format).
    """
    p = Path(source)
    if _is_xls(source):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return str(p.with_stem(f"{p.stem}_modified_{ts}").with_suffix(".xlsx"))
    if in_place:
        return source
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(p.with_stem(f"{p.stem}_modified_{ts}"))


def _save_workbook(wb: openpyxl.Workbook, path: str) -> None:
    """Save workbook with error handling."""
    try:
        wb.save(path)
        print(f"Saved: {path}")
    except Exception as exc:
        print(f"ERROR: Cannot save workbook: {exc}", file=sys.stderr)
        sys.exit(3)


def _get_sheet(wb: openpyxl.Workbook, sheet: str | None) -> Any:
    """Get worksheet by name or return active sheet."""
    if sheet is None:
        return wb.active
    if sheet not in wb.sheetnames:
        print(
            f"ERROR: Sheet '{sheet}' not found. Available: {', '.join(wb.sheetnames)}",
            file=sys.stderr,
        )
        sys.exit(2)
    return wb[sheet]


def _cell_value_str(val: Any) -> str:
    """Convert cell value to display string."""
    if val is None:
        return ""
    return str(val)


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Parse a range like 'A1:C10' into (min_col, min_row, max_col, max_row).

    Also accepts single-cell like 'B5' → (2, 5, 2, 5).
    """
    from openpyxl.utils.cell import range_boundaries

    return range_boundaries(range_str)


def _format_as_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    """Format data as a markdown table."""
    if not headers:
        return "(empty)"
    col_widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            if i < len(col_widths):
                col_widths[i] = max(col_widths[i], len(cell))

    header_line = "| " + " | ".join(h.ljust(w) for h, w in zip(headers, col_widths)) + " |"
    sep_line = "| " + " | ".join("-" * w for w in col_widths) + " |"
    data_lines = []
    for row in rows:
        padded = []
        for i, w in enumerate(col_widths):
            val = row[i] if i < len(row) else ""
            padded.append(val.ljust(w))
        data_lines.append("| " + " | ".join(padded) + " |")

    return "\n".join([header_line, sep_line, *data_lines])


def _col_label(col_idx: int) -> str:
    """1-based column index to Excel letter (1->A, 27->AA)."""
    return get_column_letter(col_idx)


def _parse_col(col_str: str) -> int:
    """Parse column as letter ('A') or 1-based number ('1') -> 1-based int."""
    try:
        return int(col_str)
    except ValueError:
        return column_index_from_string(col_str.upper())


def _effective_dimensions(ws: Any) -> tuple[int, int]:
    """Find the actual last row and column with data, ignoring trailing empties.

    openpyxl's max_row/max_column often overcount because formatting or
    metadata extends beyond real data (e.g. 256 columns in legacy .xls-origin
    files). This scans backward to find the true boundary.
    """
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    eff_col = 0
    for r in range(1, max_row + 1):
        for c in range(max_col, 0, -1):
            if ws.cell(r, c).value is not None:
                eff_col = max(eff_col, c)
                break

    eff_row = 0
    for r in range(max_row, 0, -1):
        for c in range(1, (eff_col or max_col) + 1):
            if ws.cell(r, c).value is not None:
                eff_row = r
                break
        if eff_row:
            break

    return (eff_row or max_row, eff_col or max_col)


def _auto_convert(value: str) -> Any:
    """Convert string value to appropriate Python type for Excel."""
    if value == "":
        return None
    if value.lower() in ("true", "false"):
        return value.lower() == "true"
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    if value.startswith("="):
        return value
    return value


def _bool_arg(v: str) -> bool:
    """Parse boolean argument."""
    if v.lower() in ("true", "1", "yes"):
        return True
    if v.lower() in ("false", "0", "no"):
        return False
    raise argparse.ArgumentTypeError(f"Boolean value expected, got '{v}'")


# ---------------------------------------------------------------------------
# Subcommands — read-only
# ---------------------------------------------------------------------------


def cmd_info(args: argparse.Namespace) -> None:
    """Show workbook metadata: sheets, dimensions, row/column counts."""
    wb = _load_workbook(args.file, data_only=True)
    print(f"# Workbook: {args.file}\n")
    print(f"**Sheets:** {len(wb.sheetnames)}\n")
    for name in wb.sheetnames:
        ws = wb[name]
        active_marker = " *(active)*" if ws == wb.active else ""
        eff_row, eff_col = _effective_dimensions(ws)
        print(f"## {name}{active_marker}\n")
        print(f"- Rows: {eff_row}")
        print(f"- Columns: {eff_col}")
        if eff_col > 0:
            col_labels = [_col_label(i) for i in range(1, eff_col + 1)]
            header_vals = [_cell_value_str(ws.cell(1, i).value) for i in range(1, eff_col + 1)]
            cols_with_data = [
                f"  {lbl}: {hv}" if hv else f"  {lbl}"
                for lbl, hv in zip(col_labels, header_vals)
            ]
            print(f"- Column headers (row 1):\n" + "\n".join(cols_with_data))
        if ws.merged_cells.ranges:
            print(f"- Merged cells: {', '.join(str(r) for r in ws.merged_cells.ranges)}")
        print()


def cmd_read(args: argparse.Namespace) -> None:
    """Read cell range and output as markdown table or JSON."""
    wb = _load_workbook(args.file, data_only=True)
    ws = _get_sheet(wb, args.sheet)

    if args.range:
        min_col, min_row, max_col, max_row = _parse_range(args.range)
        if max_row is None:
            max_row = ws.max_row
        if max_col is None:
            max_col = ws.max_column
    else:
        eff_row, eff_col = _effective_dimensions(ws)
        min_row = args.start_row or 1
        max_row = args.end_row or eff_row
        min_col = 1
        max_col = eff_col

    if args.limit and max_row - min_row + 1 > args.limit:
        max_row = min_row + args.limit - 1

    rows_data: list[list[str]] = []
    for r in range(min_row, max_row + 1):
        row = [_cell_value_str(ws.cell(r, c).value) for c in range(min_col, max_col + 1)]
        rows_data.append(row)

    if args.format == "json":
        if args.header and rows_data:
            headers = rows_data[0]
            records = []
            for row in rows_data[1:]:
                record = {}
                for i, h in enumerate(headers):
                    record[h if h else f"col_{i+1}"] = row[i] if i < len(row) else ""
                records.append(record)
            print(json.dumps(records, ensure_ascii=False, indent=2))
        else:
            print(json.dumps(rows_data, ensure_ascii=False, indent=2))
    else:
        col_headers = [_col_label(c) for c in range(min_col, max_col + 1)]
        if args.header and rows_data:
            headers = rows_data[0]
            table_rows = rows_data[1:]
        else:
            headers = col_headers
            table_rows = rows_data
        sheet_label = args.sheet or ws.title
        print(f"## {sheet_label} [{min_row}:{max_row}]\n")
        print(_format_as_markdown_table(headers, table_rows))
        total = max_row - min_row + 1
        if args.header:
            total -= 1
        print(f"\n({total} data rows)")


def cmd_search(args: argparse.Namespace) -> None:
    """Search for a value across sheets."""
    wb = _load_workbook(args.file, data_only=True)
    query = args.query.lower() if not args.exact else args.query
    sheets = [args.sheet] if args.sheet else wb.sheetnames
    found = 0

    print(f"# Search: \"{args.query}\" in {args.file}\n")
    for sname in sheets:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        eff_row, eff_col = _effective_dimensions(ws)
        for row in ws.iter_rows(min_row=1, max_row=eff_row, min_col=1, max_col=eff_col):
            for cell in row:
                val = _cell_value_str(cell.value)
                match = (val == args.query) if args.exact else (query in val.lower())
                if match:
                    if found == 0:
                        print("| Sheet | Cell | Value |")
                        print("| ----- | ---- | ----- |")
                    coord = f"{_col_label(cell.column)}{cell.row}"
                    print(f"| {sname} | {coord} | {val} |")
                    found += 1
                    if args.limit and found >= args.limit:
                        print(f"\n(showing first {args.limit} matches)")
                        return
    if found == 0:
        print("No matches found.")
    else:
        print(f"\n({found} matches)")


def cmd_export_csv(args: argparse.Namespace) -> None:
    """Export a sheet to CSV."""
    import csv

    wb = _load_workbook(args.file, data_only=True)
    ws = _get_sheet(wb, args.sheet)

    output_path = args.output or Path(args.file).with_suffix(".csv")
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([_cell_value_str(v) for v in row])
    print(f"Exported to: {output_path}")


# ---------------------------------------------------------------------------
# Subcommands — write (all use --in-place / safe-copy default)
# ---------------------------------------------------------------------------


def cmd_write(args: argparse.Namespace) -> None:
    """Write values to cells."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    for spec in args.cells:
        if "=" not in spec:
            print(f"ERROR: Invalid cell spec '{spec}'. Use CELL=VALUE (e.g. A1=hello).", file=sys.stderr)
            sys.exit(2)
        cell_ref, value = spec.split("=", 1)
        cell_ref = cell_ref.strip().upper()

        converted = _auto_convert(value)
        ws[cell_ref] = converted
        print(f"  {ws.title}!{cell_ref} = {value}")

    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_write_range(args: argparse.Namespace) -> None:
    """Write a block of data from JSON input."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as exc:
        print(f"ERROR: Invalid JSON: {exc}", file=sys.stderr)
        sys.exit(2)

    start_row = args.start_row or 1
    start_col = _parse_col(args.start_col) if args.start_col else 1

    if isinstance(data, list) and all(isinstance(r, dict) for r in data):
        headers = list(data[0].keys()) if data else []
        if args.header:
            for ci, h in enumerate(headers):
                ws.cell(start_row, start_col + ci, value=h)
            start_row += 1
        for ri, record in enumerate(data):
            for ci, h in enumerate(headers):
                ws.cell(start_row + ri, start_col + ci, value=record.get(h))
        count = len(data)
    elif isinstance(data, list) and all(isinstance(r, list) for r in data):
        for ri, row in enumerate(data):
            for ci, val in enumerate(row):
                ws.cell(start_row + ri, start_col + ci, value=val)
        count = len(data)
    else:
        print("ERROR: JSON must be a list of lists or a list of objects.", file=sys.stderr)
        sys.exit(2)

    print(f"Wrote {count} rows starting at {_col_label(start_col)}{args.start_row or 1}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_create(args: argparse.Namespace) -> None:
    """Create a new empty workbook."""
    p = Path(args.file)
    if p.exists() and not args.force:
        print(f"ERROR: File already exists: {args.file}. Use --force to overwrite.", file=sys.stderr)
        sys.exit(2)
    wb = openpyxl.Workbook()
    if args.sheets:
        wb.active.title = args.sheets[0]
        for name in args.sheets[1:]:
            wb.create_sheet(name)
    _save_workbook(wb, args.file)


def cmd_add_sheet(args: argparse.Namespace) -> None:
    """Add a new sheet to the workbook."""
    wb = _load_workbook(args.file)
    if args.name in wb.sheetnames:
        print(f"ERROR: Sheet '{args.name}' already exists.", file=sys.stderr)
        sys.exit(2)
    wb.create_sheet(args.name, args.position)
    print(f"Added sheet: {args.name}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_delete_sheet(args: argparse.Namespace) -> None:
    """Delete a sheet from the workbook."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.name)
    if len(wb.sheetnames) == 1:
        print("ERROR: Cannot delete the only sheet in a workbook.", file=sys.stderr)
        sys.exit(2)
    wb.remove(ws)
    print(f"Deleted sheet: {args.name}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_rename_sheet(args: argparse.Namespace) -> None:
    """Rename a sheet."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.old_name)
    if args.new_name in wb.sheetnames:
        print(f"ERROR: Sheet '{args.new_name}' already exists.", file=sys.stderr)
        sys.exit(2)
    ws.title = args.new_name
    print(f"Renamed sheet: {args.old_name} -> {args.new_name}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_copy_sheet(args: argparse.Namespace) -> None:
    """Copy a sheet within the workbook."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.source)
    new_ws = wb.copy_worksheet(ws)
    if args.name:
        new_ws.title = args.name
    print(f"Copied sheet: {args.source} -> {new_ws.title}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_add_column(args: argparse.Namespace) -> None:
    """Add a column with a header and optional default value."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    if args.position:
        col = _parse_col(args.position)
        ws.insert_cols(col)
    else:
        col = (ws.max_column or 0) + 1

    ws.cell(1, col, value=args.header)
    if args.default is not None:
        converted = _auto_convert(args.default)
        for r in range(2, (ws.max_row or 1) + 1):
            ws.cell(r, col, value=converted)
    print(f"Added column {_col_label(col)}: {args.header}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_delete_column(args: argparse.Namespace) -> None:
    """Delete a column by letter or index."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    col = _parse_col(args.column)
    header = _cell_value_str(ws.cell(1, col).value)
    ws.delete_cols(col)
    print(f"Deleted column {args.column} (was: {header})")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_rename_column(args: argparse.Namespace) -> None:
    """Rename a column header."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    col = _parse_col(args.column)
    old_header = _cell_value_str(ws.cell(1, col).value)
    ws.cell(1, col, value=args.new_name)
    print(f"Renamed column {args.column}: {old_header} -> {args.new_name}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_add_row(args: argparse.Namespace) -> None:
    """Add a row with values."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    if args.position:
        row_idx = args.position
        ws.insert_rows(row_idx)
    else:
        row_idx = (ws.max_row or 0) + 1

    if args.values:
        for ci, val in enumerate(args.values):
            ws.cell(row_idx, ci + 1, value=_auto_convert(val))
    elif args.json:
        try:
            data = json.loads(args.json)
        except json.JSONDecodeError as exc:
            print(f"ERROR: Invalid JSON: {exc}", file=sys.stderr)
            sys.exit(2)
        if isinstance(data, dict):
            headers = [_cell_value_str(ws.cell(1, c).value) for c in range(1, (ws.max_column or 0) + 1)]
            for ci, h in enumerate(headers):
                if h in data:
                    ws.cell(row_idx, ci + 1, value=data[h])
        elif isinstance(data, list):
            for ci, val in enumerate(data):
                ws.cell(row_idx, ci + 1, value=val)

    print(f"Added row at position {row_idx}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_delete_row(args: argparse.Namespace) -> None:
    """Delete one or more rows."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    rows_sorted = sorted(args.rows, reverse=True)
    for r in rows_sorted:
        ws.delete_rows(r)
    print(f"Deleted row(s): {', '.join(str(r) for r in args.rows)}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_set_style(args: argparse.Namespace) -> None:
    """Apply styling to a cell range."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)

    min_col, min_row, max_col, max_row = _parse_range(args.range)
    if max_row is None:
        max_row = min_row
    if max_col is None:
        max_col = min_col

    font_kwargs: dict[str, Any] = {}
    if args.bold is not None:
        font_kwargs["bold"] = args.bold
    if args.italic is not None:
        font_kwargs["italic"] = args.italic
    if args.font_size:
        font_kwargs["size"] = args.font_size
    if args.font_color:
        font_kwargs["color"] = args.font_color.lstrip("#")
    if args.font_name:
        font_kwargs["name"] = args.font_name

    fill = None
    if args.bg_color:
        fill = PatternFill(start_color=args.bg_color.lstrip("#"), fill_type="solid")

    alignment = None
    if args.align:
        alignment = Alignment(horizontal=args.align)

    border = None
    if args.border:
        side = Side(style=args.border)
        border = Border(left=side, right=side, top=side, bottom=side)

    font = Font(**font_kwargs) if font_kwargs else None

    count = 0
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            count += 1

    print(f"Styled {count} cells in range {args.range}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_merge(args: argparse.Namespace) -> None:
    """Merge a range of cells."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    ws.merge_cells(args.range)
    print(f"Merged cells: {args.range}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_unmerge(args: argparse.Namespace) -> None:
    """Unmerge a range of cells."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    ws.unmerge_cells(args.range)
    print(f"Unmerged cells: {args.range}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_freeze(args: argparse.Namespace) -> None:
    """Freeze panes at a given cell (e.g. A2 freezes row 1)."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    ws.freeze_panes = args.cell if args.cell.upper() != "NONE" else None
    action = "Unfroze all panes" if args.cell.upper() == "NONE" else f"Froze panes at {args.cell}"
    print(action)
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_autofilter(args: argparse.Namespace) -> None:
    """Set auto-filter on a range (e.g. A1:D100)."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    ws.auto_filter.ref = args.range
    print(f"Auto-filter set on {args.range}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_column_width(args: argparse.Namespace) -> None:
    """Set column width."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    col_letter = _col_label(_parse_col(args.column))
    ws.column_dimensions[col_letter].width = args.width
    print(f"Column {col_letter} width set to {args.width}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_row_height(args: argparse.Namespace) -> None:
    """Set row height."""
    wb = _load_workbook(args.file)
    ws = _get_sheet(wb, args.sheet)
    ws.row_dimensions[args.row].height = args.height
    print(f"Row {args.row} height set to {args.height}")
    out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)


def cmd_import_csv(args: argparse.Namespace) -> None:
    """Import CSV data into a sheet."""
    import csv

    csv_path = Path(args.csv_file)
    if not csv_path.exists():
        print(f"ERROR: CSV file not found: {args.csv_file}", file=sys.stderr)
        sys.exit(2)

    xlsx_path = Path(args.file)
    if xlsx_path.exists():
        wb = _load_workbook(args.file)
        is_new = False
    else:
        wb = openpyxl.Workbook()
        is_new = True

    sheet_name = args.sheet or csv_path.stem
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    with open(csv_path, newline="", encoding=args.encoding) as f:
        reader = csv.reader(f, delimiter=args.delimiter)
        for ri, row in enumerate(reader, start=1):
            for ci, val in enumerate(row, start=1):
                ws.cell(ri, ci, value=_auto_convert(val))

    if is_new:
        out = args.file
    else:
        out = _output_path(args.file, args.in_place)
    _save_workbook(wb, out)
    print(f"Imported {csv_path.name} -> sheet '{ws.title}'")


# ---------------------------------------------------------------------------
# CLI definition
# ---------------------------------------------------------------------------


_WRITE_COMMANDS = frozenset({
    "write", "write-range",
    "add-sheet", "delete-sheet", "rename-sheet", "copy-sheet",
    "add-column", "delete-column", "rename-column",
    "add-row", "delete-row",
    "set-style", "merge", "unmerge",
    "freeze", "autofilter", "column-width", "row-height",
    "import-csv",
})


def _add_in_place_flag(p: argparse.ArgumentParser) -> None:
    """Add the --in-place flag to a write subcommand."""
    p.add_argument(
        "--in-place", "-i",
        action="store_true",
        default=False,
        help="Modify the file in place (default: save to a timestamped copy)",
    )


def build_parser() -> argparse.ArgumentParser:
    """Build the argument parser with all subcommands."""
    parser = argparse.ArgumentParser(
        prog="excel_ops",
        description="Excel workbook operations for AI coding assistants",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    # --- info ---
    p = sub.add_parser("info", help="Show workbook metadata")
    p.add_argument("file", help="Path to .xlsx/.xls file")

    # --- read ---
    p = sub.add_parser("read", help="Read cell range as markdown or JSON")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("--sheet", "-s", help="Sheet name (default: active)")
    p.add_argument("--range", "-r", help="Cell range, e.g. A1:D10")
    p.add_argument("--start-row", type=int, help="Start row (1-based)")
    p.add_argument("--end-row", type=int, help="End row (1-based)")
    p.add_argument("--limit", "-n", type=int, help="Max rows to read")
    p.add_argument("--header", action="store_true", help="First row is header")
    p.add_argument("--format", "-f", choices=["markdown", "json"], default="markdown")

    # --- search ---
    p = sub.add_parser("search", help="Search for a value")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("query", help="Search text")
    p.add_argument("--sheet", "-s", help="Limit to one sheet")
    p.add_argument("--exact", action="store_true", help="Exact match (default: contains)")
    p.add_argument("--limit", "-n", type=int, help="Max results")

    # --- write ---
    p = sub.add_parser("write", help="Write values to cells (CELL=VALUE ...)")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("cells", nargs="+", help="Cell assignments: A1=hello B2=42")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- write-range ---
    p = sub.add_parser("write-range", help="Write a JSON block of data")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("data", help='JSON string: [[...]] or [{...}]')
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--start-row", type=int, help="Start row (default: 1)")
    p.add_argument("--start-col", help="Start column letter or number (default: A)")
    p.add_argument("--header", action="store_true", help="Write dict keys as header row")
    _add_in_place_flag(p)

    # --- create ---
    p = sub.add_parser("create", help="Create a new workbook")
    p.add_argument("file", help="Path for new .xlsx file (always creates .xlsx)")
    p.add_argument("--sheets", nargs="+", help="Sheet names to create")
    p.add_argument("--force", action="store_true", help="Overwrite if exists")

    # --- add-sheet ---
    p = sub.add_parser("add-sheet", help="Add a sheet")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("name", help="New sheet name")
    p.add_argument("--position", type=int, help="Insert position (0-based)")
    _add_in_place_flag(p)

    # --- delete-sheet ---
    p = sub.add_parser("delete-sheet", help="Delete a sheet")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("name", help="Sheet name to delete")
    _add_in_place_flag(p)

    # --- rename-sheet ---
    p = sub.add_parser("rename-sheet", help="Rename a sheet")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("old_name", help="Current name")
    p.add_argument("new_name", help="New name")
    _add_in_place_flag(p)

    # --- copy-sheet ---
    p = sub.add_parser("copy-sheet", help="Copy a sheet")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("source", help="Source sheet name")
    p.add_argument("--name", help="Name for the copy")
    _add_in_place_flag(p)

    # --- add-column ---
    p = sub.add_parser("add-column", help="Add a column")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("header", help="Column header text")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--position", help="Insert at column (letter or number)")
    p.add_argument("--default", help="Default value for existing rows")
    _add_in_place_flag(p)

    # --- delete-column ---
    p = sub.add_parser("delete-column", help="Delete a column")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("column", help="Column letter or number")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- rename-column ---
    p = sub.add_parser("rename-column", help="Rename a column header")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("column", help="Column letter or number")
    p.add_argument("new_name", help="New header text")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- add-row ---
    p = sub.add_parser("add-row", help="Add a row")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--position", type=int, help="Insert at row (1-based)")
    p.add_argument("--values", nargs="+", help="Cell values in order")
    p.add_argument("--json", help='JSON object or array of values')
    _add_in_place_flag(p)

    # --- delete-row ---
    p = sub.add_parser("delete-row", help="Delete row(s)")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("rows", nargs="+", type=int, help="Row number(s) to delete")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- set-style ---
    p = sub.add_parser("set-style", help="Style a cell range")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("range", help="Cell range, e.g. A1:D1")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--bold", type=_bool_arg, help="Bold (true/false)")
    p.add_argument("--italic", type=_bool_arg, help="Italic (true/false)")
    p.add_argument("--font-size", type=int, help="Font size")
    p.add_argument("--font-color", help="Font color hex, e.g. #FF0000")
    p.add_argument("--font-name", help="Font name, e.g. Arial")
    p.add_argument("--bg-color", help="Background color hex")
    p.add_argument("--align", choices=["left", "center", "right"])
    p.add_argument("--border", choices=["thin", "medium", "thick", "double"])
    _add_in_place_flag(p)

    # --- merge ---
    p = sub.add_parser("merge", help="Merge cells")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("range", help="Cell range to merge, e.g. A1:C1")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- unmerge ---
    p = sub.add_parser("unmerge", help="Unmerge cells")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("range", help="Cell range to unmerge")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- freeze ---
    p = sub.add_parser("freeze", help="Freeze panes")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("cell", help="Cell to freeze at (e.g. A2) or NONE to unfreeze")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- autofilter ---
    p = sub.add_parser("autofilter", help="Set auto-filter")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("range", help="Filter range, e.g. A1:D100")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- column-width ---
    p = sub.add_parser("column-width", help="Set column width")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("column", help="Column letter or number")
    p.add_argument("width", type=float, help="Width in characters")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- row-height ---
    p = sub.add_parser("row-height", help="Set row height")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("row", type=int, help="Row number")
    p.add_argument("height", type=float, help="Height in points")
    p.add_argument("--sheet", "-s", help="Sheet name")
    _add_in_place_flag(p)

    # --- export-csv ---
    p = sub.add_parser("export-csv", help="Export sheet to CSV")
    p.add_argument("file", help="Path to .xlsx/.xls file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--output", "-o", help="Output CSV path")

    # --- import-csv ---
    p = sub.add_parser("import-csv", help="Import CSV into workbook")
    p.add_argument("file", help="Path to .xlsx file (created if missing)")
    p.add_argument("csv_file", help="Path to CSV file")
    p.add_argument("--sheet", "-s", help="Target sheet name (default: CSV filename)")
    p.add_argument("--encoding", default="utf-8", help="CSV encoding (default: utf-8)")
    p.add_argument("--delimiter", default=",", help="CSV delimiter (default: ,)")
    _add_in_place_flag(p)

    return parser


COMMAND_DISPATCH = {
    "info": cmd_info,
    "read": cmd_read,
    "search": cmd_search,
    "write": cmd_write,
    "write-range": cmd_write_range,
    "create": cmd_create,
    "add-sheet": cmd_add_sheet,
    "delete-sheet": cmd_delete_sheet,
    "rename-sheet": cmd_rename_sheet,
    "copy-sheet": cmd_copy_sheet,
    "add-column": cmd_add_column,
    "delete-column": cmd_delete_column,
    "rename-column": cmd_rename_column,
    "add-row": cmd_add_row,
    "delete-row": cmd_delete_row,
    "set-style": cmd_set_style,
    "merge": cmd_merge,
    "unmerge": cmd_unmerge,
    "freeze": cmd_freeze,
    "autofilter": cmd_autofilter,
    "column-width": cmd_column_width,
    "row-height": cmd_row_height,
    "export-csv": cmd_export_csv,
    "import-csv": cmd_import_csv,
}


def main() -> None:
    """Entry point."""
    parser = build_parser()
    args = parser.parse_args()
    handler = COMMAND_DISPATCH.get(args.command)
    if handler is None:
        parser.print_help()
        sys.exit(1)
    handler(args)


if __name__ == "__main__":
    main()
