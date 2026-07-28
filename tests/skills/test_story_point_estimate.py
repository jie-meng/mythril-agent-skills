"""Tests for story-point-estimate skill scripts."""

from __future__ import annotations

import json
from pathlib import Path

import pytest


class TestBuildEstimatesSheet:
    """Tests that the workbook Estimates sheet is generated with correct structure."""

    @pytest.fixture(autouse=True)
    def _setup(self):
        from generate_report import _build_estimates_sheet

        self._build = _build_estimates_sheet

    def test_sheet_headers_and_data_no_platforms(self):
        from openpyxl import Workbook

        wb = Workbook()
        estimates = [
            {
                "category": "Functional",
                "epic": "Auth",
                "area": "",
                "story": "Login",
                "role": "User",
                "want": "log in",
                "so_that": "access account",
                "points": 3,
                "rationale": "Standard form",
                "uncertainty": "Low",
            },
        ]
        self._build(wb, estimates, [])
        ws = wb.active

        assert ws.cell(row=1, column=1).value == "#"
        # Column 9 = Points (8 base headers + 0 platforms = positions 1-8, then Points at 9)
        assert ws.cell(row=1, column=9).value == "Points"
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=5).value == "Login"
        assert ws.cell(row=2, column=9).value == 3
        # Uncertainty is column 11
        assert ws.cell(row=2, column=11).value == "Low"

    def test_sheet_with_platforms(self):
        from openpyxl import Workbook

        wb = Workbook()
        platforms = ["设备端", "后端", "Android", "iOS", "WEB"]
        estimates = [
            {
                "category": "Functional",
                "epic": "GPS",
                "story": "GNSS 定位",
                "role": "设备",
                "want": "定位",
                "so_that": "知道位置",
                "points": 5,
                "platform_points": {"设备端": 5, "后端": 0, "Android": 0, "iOS": 0, "WEB": 0},
                "rationale": "多模GNSS",
                "uncertainty": "Medium",
            },
        ]
        self._build(wb, estimates, platforms)
        ws = wb.active

        # Headers: #, Category, Epic, Area, Story, Role, I want..., So that...,
        #          设备端, 后端, Android, iOS, WEB, Points, Rationale, Uncertainty
        # = 8 + 5 + 3 = 16 columns
        assert ws.cell(row=1, column=1).value == "#"
        assert ws.cell(row=1, column=9).value == "设备端"
        assert ws.cell(row=1, column=10).value == "后端"
        assert ws.cell(row=1, column=13).value == "WEB"
        assert ws.cell(row=1, column=14).value == "Points"

        # Data row
        assert ws.cell(row=2, column=9).value == 5  # 设备端
        assert ws.cell(row=2, column=10).value == 0  # 后端
        assert ws.cell(row=2, column=14).value == 5  # Points
        assert ws.cell(row=2, column=16).value == "Medium"  # Uncertainty

    def test_empty_estimates(self):
        from openpyxl import Workbook

        wb = Workbook()
        self._build(wb, [], [])
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "#"
        assert ws.cell(row=2, column=1).value is None


class TestBuildPlatformSummary:
    """Tests for Platform Summary sheet."""

    @pytest.fixture(autouse=True)
    def _setup(self):
        from generate_report import _build_platform_summary

        self._build_ps = _build_platform_summary

    def test_platform_summary_structure(self):
        from openpyxl import Workbook

        wb = Workbook()
        platforms = ["设备端", "后端", "Android"]
        estimates = [
            {
                "category": "Functional",
                "platform_points": {"设备端": 5, "后端": 2, "Android": 3},
                "points": 8,
                "uncertainty": "Low",
            },
            {
                "category": "CFR",
                "platform_points": {"设备端": 0, "后端": 3, "Android": 0},
                "points": 3,
                "uncertainty": "Low",
            },
        ]
        buffers = [
            {"type": "集成联调", "pct": "10%", "points": 1, "rationale": "跨端测试"},
        ]
        self._build_ps(wb, estimates, platforms, 8, 3, 11, buffers)
        ws = wb["Platform Summary"]

        # Header row
        assert ws.cell(row=1, column=1).value == "Module"
        assert ws.cell(row=1, column=2).value == "设备端"

        # Functional row: 设备端=5, 后端=2, Android=3, subtotal=10
        assert ws.cell(row=2, column=1).value == "Functional Stories"
        assert ws.cell(row=2, column=2).value == 5
        assert ws.cell(row=2, column=5).value == 10  # subtotal

        # CFR row: 设备端=0, 后端=3, Android=0, subtotal=3
        assert ws.cell(row=3, column=1).value == "CFR Items"
        assert ws.cell(row=3, column=3).value == 3

        # "Recommended Planning Estimate" present
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Recommended" in str(cell.value):
                    found = True
        assert found

    def test_platform_summary_no_buffers(self):
        from openpyxl import Workbook

        wb = Workbook()
        platforms = ["后端", "WEB"]
        estimates = [
            {
                "category": "Functional",
                "platform_points": {"后端": 5, "WEB": 3},
                "points": 8,
            },
        ]
        self._build_ps(wb, estimates, platforms, 8, 0, 8, [])
        ws = wb["Platform Summary"]
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Buffer Total" in str(cell.value):
                    found = True
        assert not found


class TestBuildRaidSheet:
    """Tests for RAID sheet generation."""

    @pytest.fixture(autouse=True)
    def _setup(self):
        from generate_report import _build_raid_sheet

        self._build_raid = _build_raid_sheet

    def test_raid_headers_and_data(self):
        from openpyxl import Workbook

        wb = Workbook()
        raid = [
            {
                "type": "Risk",
                "item": "API downtime",
                "impact": "Service unavailable",
                "mitigation": "Circuit breaker",
            },
        ]
        self._build_raid(wb, raid)
        ws = wb["RAID"]

        assert ws.cell(row=1, column=1).value == "Type"
        assert ws.cell(row=2, column=1).value == "Risk"
        assert ws.cell(row=2, column=2).value == "API downtime"
        assert ws.cell(row=2, column=3).value == "Service unavailable"
        assert ws.cell(row=2, column=4).value == "Circuit breaker"


class TestBuildSummarySheet:
    """Tests for Summary sheet generation."""

    @pytest.fixture(autouse=True)
    def _setup(self):
        from generate_report import _build_summary_sheet

        self._build_summary = _build_summary_sheet

    def test_summary_contains_scope(self):
        from openpyxl import Workbook

        wb = Workbook()
        data = {
            "scope_summary": "Test scope",
            "key_assumptions": ["A1"],
            "recommendations": "R1",
            "confidence": "Medium",
            "estimates": [],
        }
        self._build_summary(wb, data, 10, 5, 15)
        ws = wb["Summary"]

        assert ws.cell(row=1, column=1).value == "Scope Summary"
        assert ws.cell(row=2, column=1).value == "Test scope"

    def test_totals(self):
        from openpyxl import Workbook

        wb = Workbook()
        data = {
            "scope_summary": "",
            "key_assumptions": [],
            "recommendations": "",
            "confidence": "High",
            "estimates": [],
        }
        self._build_summary(wb, data, 20, 10, 30)
        ws = wb["Summary"]

        found = False
        for row in ws.iter_rows(min_row=1, max_col=3):
            for cell in row:
                if cell.value == "Grand Total":
                    total_cell = ws.cell(row=cell.row, column=2)
                    assert total_cell.value == 30
                    found = True
        assert found

    def test_confidence(self):
        from openpyxl import Workbook

        wb = Workbook()
        data = {
            "scope_summary": "",
            "key_assumptions": [],
            "recommendations": "",
            "confidence": "Low",
            "estimates": [],
        }
        self._build_summary(wb, data, 0, 0, 0)
        ws = wb["Summary"]

        found = False
        for row in ws.iter_rows(min_row=1, max_col=3):
            for cell in row:
                if cell.value and "Low" in str(cell.value) and "Confidence" in str(cell.value):
                    found = True
        assert found
