"""Tests for excel skill scripts."""

from __future__ import annotations

import json

import pytest


class TestAutoConvert:
    """Tests for excel_ops._auto_convert."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _auto_convert
        self.func = _auto_convert

    def test_empty_string_returns_none(self):
        assert self.func("") is None

    def test_integer(self):
        assert self.func("42") == 42
        assert isinstance(self.func("42"), int)

    def test_negative_integer(self):
        assert self.func("-7") == -7

    def test_float(self):
        assert self.func("3.14") == pytest.approx(3.14)
        assert isinstance(self.func("3.14"), float)

    def test_boolean_true(self):
        assert self.func("true") is True
        assert self.func("True") is True
        assert self.func("TRUE") is True

    def test_boolean_false(self):
        assert self.func("false") is False
        assert self.func("False") is False

    def test_formula_preserved(self):
        assert self.func("=SUM(A1:A10)") == "=SUM(A1:A10)"

    def test_plain_string(self):
        assert self.func("hello world") == "hello world"

    def test_string_with_leading_zero(self):
        assert self.func("007") == 7


class TestCellValueStr:
    """Tests for excel_ops._cell_value_str."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _cell_value_str
        self.func = _cell_value_str

    def test_none_returns_empty(self):
        assert self.func(None) == ""

    def test_integer(self):
        assert self.func(42) == "42"

    def test_float(self):
        assert self.func(3.14) == "3.14"

    def test_string(self):
        assert self.func("hello") == "hello"

    def test_boolean(self):
        assert self.func(True) == "True"


class TestParseCol:
    """Tests for excel_ops._parse_col."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _parse_col
        self.func = _parse_col

    def test_letter_a(self):
        assert self.func("A") == 1

    def test_letter_z(self):
        assert self.func("Z") == 26

    def test_letter_aa(self):
        assert self.func("AA") == 27

    def test_lowercase(self):
        assert self.func("c") == 3

    def test_numeric_string(self):
        assert self.func("5") == 5


class TestColLabel:
    """Tests for excel_ops._col_label."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _col_label
        self.func = _col_label

    def test_first_column(self):
        assert self.func(1) == "A"

    def test_26th_column(self):
        assert self.func(26) == "Z"

    def test_27th_column(self):
        assert self.func(27) == "AA"


class TestFormatAsMarkdownTable:
    """Tests for excel_ops._format_as_markdown_table."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _format_as_markdown_table
        self.func = _format_as_markdown_table

    def test_empty_headers(self):
        assert self.func([], []) == "(empty)"

    def test_simple_table(self):
        result = self.func(["A", "B"], [["1", "2"], ["3", "4"]])
        lines = result.split("\n")
        assert len(lines) == 4
        assert "A" in lines[0]
        assert "---" in lines[1] or "- |" in lines[1]

    def test_uneven_rows(self):
        result = self.func(["Name", "Score"], [["Alice"]])
        assert "Alice" in result

    def test_column_width_adapts(self):
        result = self.func(["X"], [["LongValue"]])
        header_line = result.split("\n")[0]
        sep_line = result.split("\n")[1]
        assert len(header_line) == len(sep_line)


class TestBoolArg:
    """Tests for excel_ops._bool_arg."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _bool_arg
        self.func = _bool_arg

    def test_true_variants(self):
        for v in ("true", "True", "1", "yes"):
            assert self.func(v) is True

    def test_false_variants(self):
        for v in ("false", "False", "0", "no"):
            assert self.func(v) is False

    def test_invalid_raises(self):
        import argparse
        with pytest.raises(argparse.ArgumentTypeError):
            self.func("maybe")


class TestIsXls:
    """Tests for excel_ops._is_xls."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _is_xls
        self.func = _is_xls

    def test_xls_extension(self):
        assert self.func("report.xls") is True

    def test_xls_uppercase(self):
        assert self.func("REPORT.XLS") is True

    def test_xlsx_extension(self):
        assert self.func("report.xlsx") is False

    def test_xlsm_extension(self):
        assert self.func("data.xlsm") is False

    def test_no_extension(self):
        assert self.func("report") is False


class TestOutputPath:
    """Tests for excel_ops._output_path."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _output_path
        self.func = _output_path

    def test_in_place_returns_same(self):
        assert self.func("/tmp/report.xlsx", True) == "/tmp/report.xlsx"

    def test_copy_contains_modified(self):
        result = self.func("/tmp/report.xlsx", False)
        assert "report_modified_" in result
        assert result.endswith(".xlsx")

    def test_copy_preserves_directory(self):
        from pathlib import Path
        result = self.func("/some/dir/data.xlsx", False)
        assert str(Path(result).parent) == "/some/dir"

    def test_xls_source_always_outputs_xlsx(self):
        result = self.func("/tmp/report.xls", True)
        assert result.endswith(".xlsx")
        assert "modified_" in result

    def test_xls_source_never_in_place(self):
        result = self.func("/tmp/report.xls", False)
        assert result.endswith(".xlsx")
        assert "report_modified_" in result


class TestEffectiveDimensions:
    """Tests for excel_ops._effective_dimensions."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import _effective_dimensions
        self.func = _effective_dimensions

    def test_trims_trailing_empty_columns(self, tmp_path):
        import openpyxl
        p = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Name"
        ws["B1"] = "Score"
        ws["C1"] = "Grade"
        ws["A2"] = "Alice"
        ws["B2"] = 95
        # Simulate formatting on far-right column (like legacy xls)
        ws.cell(1, 100).number_format = "0.00"
        wb.save(p)

        wb2 = openpyxl.load_workbook(p, data_only=True)
        eff_row, eff_col = self.func(wb2.active)
        assert eff_col == 3
        assert eff_row == 2

    def test_empty_sheet(self, tmp_path):
        import openpyxl
        p = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        wb.save(p)

        wb2 = openpyxl.load_workbook(p, data_only=True)
        eff_row, eff_col = self.func(wb2.active)
        assert eff_row >= 0
        assert eff_col >= 0


class TestCmdCreate:
    """Tests for excel_ops.cmd_create (filesystem)."""

    @pytest.fixture(autouse=True)
    def _import(self):
        from excel_ops import cmd_create
        self.func = cmd_create

    def test_create_simple(self, tmp_path):
        import argparse
        out = tmp_path / "test.xlsx"
        args = argparse.Namespace(file=str(out), sheets=None, force=False)
        self.func(args)
        assert out.exists()

    def test_create_with_sheets(self, tmp_path):
        import argparse
        import openpyxl
        out = tmp_path / "test.xlsx"
        args = argparse.Namespace(file=str(out), sheets=["Data", "Summary"], force=False)
        self.func(args)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Data", "Summary"]

    def test_create_no_overwrite(self, tmp_path):
        import argparse
        out = tmp_path / "test.xlsx"
        out.write_bytes(b"dummy")
        args = argparse.Namespace(file=str(out), sheets=None, force=False)
        with pytest.raises(SystemExit):
            self.func(args)

    def test_create_force_overwrite(self, tmp_path):
        import argparse
        out = tmp_path / "test.xlsx"
        out.write_bytes(b"dummy")
        args = argparse.Namespace(file=str(out), sheets=None, force=True)
        self.func(args)
        assert out.exists()


class TestCmdWriteAndRead:
    """Integration tests: write then read back."""

    def test_write_in_place(self, tmp_path):
        import argparse
        import openpyxl
        from excel_ops import cmd_create, cmd_write

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=None, force=False))
        cmd_write(argparse.Namespace(
            file=path, sheet=None, in_place=True,
            cells=["A1=Name", "B1=Score", "A2=Alice", "B2=95"],
        ))

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Score"
        assert ws["A2"].value == "Alice"
        assert ws["B2"].value == 95

    def test_write_creates_copy(self, tmp_path):
        import argparse
        from pathlib import Path
        from excel_ops import cmd_create, cmd_write

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=None, force=False))
        cmd_write(argparse.Namespace(
            file=path, sheet=None, in_place=False,
            cells=["A1=Hello"],
        ))

        copies = list(tmp_path.glob("test_modified_*.xlsx"))
        assert len(copies) == 1


class TestCmdAddDeleteSheet:
    """Tests for sheet add/delete operations."""

    def test_add_and_delete_sheet(self, tmp_path):
        import argparse
        import openpyxl
        from excel_ops import cmd_create, cmd_add_sheet, cmd_delete_sheet

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=None, force=False))
        cmd_add_sheet(argparse.Namespace(file=path, name="NewSheet", position=None, in_place=True))

        wb = openpyxl.load_workbook(path)
        assert "NewSheet" in wb.sheetnames

        cmd_delete_sheet(argparse.Namespace(file=path, name="NewSheet", in_place=True))
        wb = openpyxl.load_workbook(path)
        assert "NewSheet" not in wb.sheetnames


class TestCmdRenameSheet:
    """Tests for sheet rename."""

    def test_rename(self, tmp_path):
        import argparse
        import openpyxl
        from excel_ops import cmd_create, cmd_rename_sheet

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=["Old"], force=False))
        cmd_rename_sheet(argparse.Namespace(
            file=path, old_name="Old", new_name="New", in_place=True,
        ))

        wb = openpyxl.load_workbook(path)
        assert "New" in wb.sheetnames
        assert "Old" not in wb.sheetnames


class TestCmdColumnOps:
    """Tests for column add/delete/rename."""

    def test_add_column(self, tmp_path):
        import argparse
        import openpyxl
        from excel_ops import cmd_create, cmd_write, cmd_add_column

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=None, force=False))
        cmd_write(argparse.Namespace(
            file=path, sheet=None, in_place=True,
            cells=["A1=Name", "A2=Alice"],
        ))
        cmd_add_column(argparse.Namespace(
            file=path, sheet=None, header="Score",
            position=None, default="0", in_place=True,
        ))

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(1, 2).value == "Score"
        assert ws.cell(2, 2).value == 0

    def test_rename_column(self, tmp_path):
        import argparse
        import openpyxl
        from excel_ops import cmd_create, cmd_write, cmd_rename_column

        path = str(tmp_path / "test.xlsx")
        cmd_create(argparse.Namespace(file=path, sheets=None, force=False))
        cmd_write(argparse.Namespace(
            file=path, sheet=None, in_place=True,
            cells=["A1=Old"],
        ))
        cmd_rename_column(argparse.Namespace(
            file=path, sheet=None, column="A", new_name="New", in_place=True,
        ))

        wb = openpyxl.load_workbook(path)
        assert wb.active.cell(1, 1).value == "New"
